"""
UiPath Extra Sources - project.json + Config.xlsx

Reads project-level files that scanner.py never touches (it only walks .xaml files).
These are relevant *only* to the AI-export pipeline, not to the interactive project
map: project.json gives structural/dependency context ("what kind of process is this"),
and Config.xlsx gives the process's externally-configured parameters ("what does this
process depend on that isn't visible in the XAML logic itself").

Both readers are best-effort and independent: a project without Config.xlsx (or with a
non-standard project.json) still returns a usable, partially-populated result rather
than raising - see build_extra_sources().
"""

import json
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - degrade gracefully if not installed
    openpyxl = None


# ---------------------------------------------------------------------------
# project.json
# ---------------------------------------------------------------------------

def read_project_json(project_root: Path) -> dict | None:
    """
    Reads project.json for structural context that isn't derivable from the XAML files
    themselves:
    - main: the true entry point (lets the AI know where to start reading, instead of
      guessing from invoked_by counts).
    - dependencies: the installed activity packages. This is a cheap, reliable signal
      of what *kind* of process this is (e.g. Excel + Mail packages but no
      UIAutomation package strongly suggests a data/email process, not a UI-automation
      one) - free business context that costs nothing to extract.
    - entryPoints: present in newer multi-entry-point projects (e.g. Automation Hub /
      Studio Web API processes); listed as-is when present.

    Deliberately NOT extracted: "description" and "name". In practice these are almost
    always left at their Studio-generated defaults ("Blank Process", the folder name)
    rather than genuinely describing the project, so including them would cost tokens
    while teaching the AI nothing reliable.

    Returns None if project.json doesn't exist or can't be parsed - the caller treats
    that as "no project-level metadata available", not an error.
    """
    path = project_root / "project.json"
    if not path.exists():
        return None

    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None

    result: dict = {
        "main": data.get("main"),
        "dependencies": data.get("dependencies", {}),
    }

    entry_points = data.get("entryPoints")
    if entry_points:
        result["entry_points"] = entry_points

    return result


# ---------------------------------------------------------------------------
# Config.xlsx
# ---------------------------------------------------------------------------

# Common sheet names across the standard REFramework-style Config.xlsx layout. Matched
# case-insensitively since capitalization varies by template/UiPath version.
_SETTINGS_SHEET_NAMES = {"settings"}
_ASSETS_SHEET_NAMES = {"assets"}


def _find_config_file(project_root: Path) -> Path | None:
    """
    Looks for a Config.xlsx (or Config.xls) directly under the project root or one
    level down (e.g. a Data/ or Config/ subfolder) - the two placements seen in
    practice. Returns the first match; None if no config workbook is found.
    """
    candidates = [
        project_root / "Config.xlsx",
        project_root / "Config.xls",
    ]
    for c in candidates:
        if c.exists():
            return c

    for sub in project_root.iterdir() if project_root.exists() else []:
        if sub.is_dir():
            for name in ("Config.xlsx", "Config.xls"):
                candidate = sub / name
                if candidate.exists():
                    return candidate

    return None


def _read_two_column_sheet(ws) -> list[dict]:
    """
    Reads a Name/Value-style sheet (Settings, Assets) into a list of {name, value}
    rows. Assumes the first row is a header and the first two columns are
    Name/Value - the standard layout. Skips fully empty rows. Values are stringified
    since they're headed for a text prompt, not further computation.
    """
    rows = []
    row_iter = ws.iter_rows(min_row=2, values_only=True)
    for row in row_iter:
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name:
            continue
        value = row[1] if len(row) > 1 else None
        rows.append({"name": name, "value": "" if value is None else str(value)})
    return rows


def read_config_workbook(project_root: Path) -> dict | None:
    """
    Reads Config.xlsx and extracts Settings/Assets as name/value pairs. Only sheet
    *names* and *keys* are meaningfully structured here - values are passed through
    as-is, including Asset values, which are typically references (queue names,
    credential identifiers) rather than secrets themselves. If a project stores actual
    secrets in Config.xlsx (non-standard practice), the caller is responsible for not
    forwarding this data to an AI - see build_extra_sources() docstring.

    Returns None if no config workbook is found or it can't be read. Any sheet whose
    name doesn't match the known Settings/Assets patterns is ignored, since its
    structure isn't assumed to be Name/Value and free-form parsing risks pulling in
    noise.
    """
    if openpyxl is None:
        return None

    path = _find_config_file(project_root)
    if path is None:
        return None

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None

    result: dict = {"source_file": path.name, "settings": [], "assets": [], "other_sheets": []}

    for sheet_name in wb.sheetnames:
        normalized = sheet_name.strip().lower()
        ws = wb[sheet_name]
        if normalized in _SETTINGS_SHEET_NAMES:
            result["settings"] = _read_two_column_sheet(ws)
        elif normalized in _ASSETS_SHEET_NAMES:
            result["assets"] = _read_two_column_sheet(ws)
        else:
            result["other_sheets"].append(sheet_name)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Combined entry point
# ---------------------------------------------------------------------------

def build_extra_sources(project_root: Path) -> dict:
    """
    Runs both readers and returns a single dict for ai_export.py to fold in alongside
    the XAML catalog. Each sub-result is None if that source wasn't found/readable -
    ai_export.py is expected to omit a key entirely from the final prompt when its
    value is None/empty, rather than emitting a "not found" placeholder that would
    just cost tokens for no benefit.

    NOTE on sensitive data: Config.xlsx's Assets sheet is included because Asset
    *names* (e.g. "SharedMailbox", "SAP_Credentials") are useful business-context
    signals, and Orchestrator Assets don't store their actual secret values in this
    file to begin with - only a reference name. Settings values are passed through
    unfiltered; if a given project's Config.xlsx contains anything sensitive in
    Settings (uncommon but possible), that's a call for the UI layer to let the user
    review before copying to an AI chat, not something this module tries to guess at.
    """
    return {
        "project_json": read_project_json(project_root),
        "config": read_config_workbook(project_root),
    }


if __name__ == "__main__":
    import sys

    root = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
    result = build_extra_sources(root)
    print(json.dumps(result, indent=2, ensure_ascii=False))